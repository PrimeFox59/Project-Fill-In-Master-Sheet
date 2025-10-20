
import streamlit as st
import pandas as pd
import io
import openpyxl
import sqlite3
import json
from pathlib import Path
from datetime import datetime

# Optional viz libs
try:
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except Exception:  # pragma: no cover
    px = None
    PLOTLY_AVAILABLE = False

# Fungsi untuk memproses data (tidak ada perubahan di fungsi ini)
def process_data(source_df, template_df, target_column, mode):
    """
    Fungsi ini mengambil data sumber, menghitungnya, dan memperbarui DataFrame template.
    """
    # 1. Ekstrak dan hitung data dari file sumber
    # Filter baris di mana 'Student Code' dan 'Course Code' tidak kosong
    filtered_df = source_df.dropna(subset=['Student Code', 'Course Code'])
    # Hitung kemunculan setiap 'Site Name'
    site_counts = filtered_df['Site Name'].value_counts()
    # Buat salinan template_df agar tidak mengubah data asli secara langsung
    updated_df = template_df.copy()
    # Asumsikan kolom A (kolom pertama) di template adalah untuk 'Site Name'
    site_name_col_in_template = updated_df.columns[0]
    # 2. Iterasi melalui hasil hitungan dan perbarui template
    for site_name, count in site_counts.items():
        # Cari baris yang cocok dengan site_name di template
        match_row = updated_df[updated_df[site_name_col_in_template] == site_name]
        if not match_row.empty:
            # Jika Site Name ditemukan
            idx = match_row.index[0]
            if mode == "Add (Tambah)":
                # Ambil nilai saat ini, ubah ke numerik, anggap 0 jika kosong/error
                current_value = pd.to_numeric(updated_df.loc[idx, target_column], errors='coerce').fillna(0)
                updated_df.loc[idx, target_column] = current_value + count
            else:  # Mode "Replace (Ganti)"
                updated_df.loc[idx, target_column] = count
        else:
            # Jika Site Name tidak ditemukan
            # Buat baris baru sebagai dictionary
            new_row = {site_name_col_in_template: site_name, target_column: count}
            # Tambahkan baris baru ke DataFrame
            updated_df = pd.concat([updated_df, pd.DataFrame([new_row])], ignore_index=True)
    return updated_df

# --- SQLite helpers ---
DB_PATH = Path(__file__).with_name("master_sheet.db")

def _get_conn():
    # check_same_thread=False allows usage across Streamlit threads
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    with _get_conn() as conn:
        c = conn.cursor()
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                site_name_header TEXT NOT NULL,
                columns_json TEXT NOT NULL
            )
            """
        )
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS master_values (
                run_id INTEGER NOT NULL,
                site_name TEXT NOT NULL,
                category TEXT NOT NULL,
                value REAL,
                PRIMARY KEY (run_id, site_name, category),
                FOREIGN KEY(run_id) REFERENCES runs(id) ON DELETE CASCADE
            )
            """
        )
        conn.commit()

def save_result_to_db(result_df: pd.DataFrame, header: list[str]) -> dict:
    """Persist a processed wide table into SQLite in a normalized form.
    Returns metadata for the run: {run_id, created_at, site_name_header, categories}
    """
    if not isinstance(result_df, pd.DataFrame) or not header or len(header) < 1:
        return {}
    init_db()
    site_name_header = str(header[0])
    categories = [str(c) for c in header[1:]]
    created_at = datetime.utcnow().isoformat()

    with _get_conn() as conn:
        c = conn.cursor()
        c.execute(
            "INSERT INTO runs (created_at, site_name_header, columns_json) VALUES (?, ?, ?)",
            (created_at, site_name_header, json.dumps(categories)),
        )
        run_id = c.lastrowid

        # Prepare insert
        rows_to_insert = []
        for _, row in result_df.iterrows():
            site_name = row.get(site_name_header, None)
            if pd.isna(site_name):
                continue
            site_name = str(site_name)
            for cat in categories:
                val = row.get(cat, None)
                # Coerce to numeric if possible, otherwise NULL
                try:
                    val_num = pd.to_numeric(val, errors="coerce")
                    val_out = None if pd.isna(val_num) else float(val_num)
                except Exception:
                    val_out = None
                rows_to_insert.append((run_id, site_name, cat, val_out))

        c.executemany(
            "INSERT OR REPLACE INTO master_values (run_id, site_name, category, value) VALUES (?, ?, ?, ?)",
            rows_to_insert,
        )
        conn.commit()

    return {"run_id": run_id, "created_at": created_at, "site_name_header": site_name_header, "categories": categories}

def load_latest_from_db() -> dict | None:
    """Load the most recent run and reconstruct a wide DataFrame.
    Returns dict: { df, meta }
    """
    if not DB_PATH.exists():
        return None
    init_db()
    with _get_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT id, created_at, site_name_header, columns_json FROM runs ORDER BY id DESC LIMIT 1")
        row = c.fetchone()
        if not row:
            return None
        run_id, created_at, site_name_header, columns_json = row
        categories = json.loads(columns_json)
        c.execute("SELECT site_name, category, value FROM master_values WHERE run_id = ?", (run_id,))
        vals = c.fetchall()

    # Build wide DataFrame
    data_map = {}
    for site_name, category, value in vals:
        if site_name not in data_map:
            data_map[site_name] = {cat: None for cat in categories}
        if category in data_map[site_name]:
            data_map[site_name][category] = value

    rows = []
    for site, cat_map in data_map.items():
        row = {site_name_header: site}
        row.update(cat_map)
        rows.append(row)

    if not rows:
        return None
    df = pd.DataFrame(rows)
    # Ensure column order: site name first, then categories
    df = df[[site_name_header] + categories]
    meta = {"run_id": run_id, "created_at": created_at, "site_name_header": site_name_header, "categories": categories}
    return {"df": df, "meta": meta}

# --- UI Streamlit ---
st.set_page_config(layout="wide")

# --- Constants ---
# Categories are derived dynamically from the 'Master Sheet' header (excluding the first column).

# --- App Title & Sidebar ---
st.title("📊 Master Sheet Assistant")
st.caption("Separate menus: Dashboard & Data Input")

menu = st.sidebar.radio("Menu", ["Dashboard", "Data Input", "User Guide"], index=1)

# Session placeholders
if "template_df" not in st.session_state:
    st.session_state.template_df = None
if "result_df" not in st.session_state:
    st.session_state.result_df = None
if "last_processed_ext" not in st.session_state:
    st.session_state.last_processed_ext = ".xlsx"

def render_dashboard():
    st.header("📈 Partner Engagement Dashboard")
    st.write("Explore the Master Sheet data interactively: summary, Top/Bottom, category comparisons, and company profile.")

    # Simple styling touch
    st.markdown(
        """
        <style>
                .section-title {margin-top: 0.5rem;}

                /* Modern KPI cards */
                .kpi-card{
                    position: relative;
                    display: flex; flex-direction: column; gap: 6px;
                    padding: 14px 16px; border-radius: 14px;
                    background: linear-gradient(135deg, #ffffff 0%, #f5f8ff 100%);
                    border: 1px solid #e6eaf2; box-shadow: 0 2px 10px rgba(28, 39, 60, 0.06);
                    min-height: 110px;
                }
                .kpi-head{display:flex; align-items:center; gap:8px; color:#4b5563; font-weight:600; font-size: 0.95rem;}
                .kpi-icon{width: 36px; height: 36px; border-radius: 10px; display:flex; align-items:center; justify-content:center; color:#fff; font-size: 18px;}
                .kpi-icon.primary{background: linear-gradient(135deg, #2563eb, #1d4ed8);} /* blue */
                .kpi-icon.success{background: linear-gradient(135deg, #10b981, #059669);}  /* green */
                .kpi-icon.warning{background: linear-gradient(135deg, #f59e0b, #d97706);}  /* orange */
                .kpi-icon.slate{background: linear-gradient(135deg, #64748b, #475569);}    /* slate */
                .kpi-value{font-size: 2rem; font-weight: 700; letter-spacing: -0.5px; color:#111827; margin-top:4px;}
                .kpi-sub{color:#6b7280; font-size: 0.85rem;}
                .kpi-badge{position:absolute; right:10px; top:10px; font-size: 0.75rem; padding: 2px 8px; border-radius: 999px; background:#eef2ff; color:#3730a3; border:1px solid #e0e7ff;}
                .muted{color:#6b7280}
        
                /* Subtle animated gradient dot */
                .kpi-dot{position:absolute; right:14px; bottom:14px; width:8px; height:8px; border-radius:999px; background: radial-gradient(circle at 30% 30%, #93c5fd, #2563eb); opacity:0.5}
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Try loading latest processed data from DB if session template is empty
    template_df = st.session_state.template_df
    if template_df is None:
        loaded = load_latest_from_db()
        if loaded is not None:
            template_df = loaded["df"]
            st.session_state.template_df = template_df
            st.session_state._db_meta = loaded["meta"]
            st.info(
                f"Using latest processed data from database (run at {loaded['meta']['created_at']} UTC). Upload a template to override.")

    # Upload template if not already in session
    uploaded_template = st.file_uploader(
        "Upload Excel template file (must contain sheet 'Master Sheet')",
        type=["xlsx", "xlsm"],
        key="dashboard_template_uploader",
    )

    template_df = st.session_state.template_df
    if uploaded_template is not None:
        try:
            template_df = pd.read_excel(uploaded_template, sheet_name="Master Sheet")
            st.session_state.template_df = template_df
        except Exception as e:
            st.error(f"Failed to read template: {e}")
            return

    if template_df is None:
        st.info("Please upload a template here or in the 'Data Input' menu first.")
        return

    # Validasi kategori yang ada di template
    site_name_col = template_df.columns[0]
    available_categories = [c for c in template_df.columns if c != site_name_col]
    if not available_categories:
        st.warning("No category columns found in 'Master Sheet' besides the first column (company name).")
        st.write("Available headers in the template:")
        st.code("\n".join(list(map(str, template_df.columns))))
        return
    # Helpers
    def as_numeric(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce").fillna(0)

    # Tabs for different perspectives
    tab_overview, tab_top, tab_matrix, tab_profile = st.tabs([
        "Overview", "Top/Bottom", "Matrix", "Company Profile"
    ])

    with tab_overview:
        c1, c2 = st.columns([2, 1])
        with c2:
            category = st.selectbox("Select Category", options=available_categories, key="ov_cat")
        with c1:
            st.subheader("Category Summary", anchor=False)

        df_cat = template_df[[site_name_col, category]].copy()
        df_cat[category] = as_numeric(df_cat[category])
        df_cat = df_cat.dropna(subset=[site_name_col])
        total_val = float(df_cat[category].sum())
        avg_val = float(df_cat[category].mean())
        max_row = df_cat.loc[df_cat[category].idxmax()] if not df_cat.empty else None
        min_row = df_cat.loc[df_cat[category].idxmin()] if not df_cat.empty else None

        def fmt_value(v: float) -> str:
            if v is None:
                return "-"
            # show two decimals for small numbers, else thousands separator
            if abs(v) < 1:
                return f"{v:,.2f}"
            # if integer-like, no decimals
            if float(v).is_integer():
                return f"{int(v):,}"
            return f"{v:,.2f}"

        def metric_card(title: str, value: float, subtitle: str = "", icon: str = "📊", tone: str = "primary"):
            html = f"""
            <div class='kpi-card'>
              <div class='kpi-head'>
                 <div class='kpi-icon {tone}'>{icon}</div>
                 <div>{title}</div>
                 <div class='kpi-badge'>Category</div>
              </div>
              <div class='kpi-value'>{fmt_value(value)}</div>
              <div class='kpi-sub'>{subtitle}</div>
              <div class='kpi-dot'></div>
            </div>
            """
            st.markdown(html, unsafe_allow_html=True)

        k1, k2, k3, k4 = st.columns(4)
        with k1:
            metric_card("Total", total_val, subtitle=f"Sum of values for {category}", icon="Σ", tone="primary")
        with k2:
            metric_card("Average", avg_val, subtitle=f"Average per company", icon="⌀", tone="slate")
        with k3:
            metric_card(
                "Maximum",
                float(max_row[category]) if max_row is not None else 0,
                subtitle=f"Company: {str(max_row[site_name_col]) if max_row is not None else '-'}",
                icon="⬆",
                tone="success",
            )
        with k4:
            metric_card(
                "Minimum",
                float(min_row[category]) if min_row is not None else 0,
                subtitle=f"Company: {str(min_row[site_name_col]) if min_row is not None else '-'}",
                icon="⬇",
                tone="warning",
            )

        st.markdown("### Value Distribution", help="Distribution of values in the selected category.")
        if PLOTLY_AVAILABLE:
            fig = px.histogram(df_cat, x=category, nbins=20, title="Histogram", template="simple_white")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.bar_chart(df_cat[category])

        st.markdown("### Sample Data", help="Top 5 and Bottom 5 values.")
        colA, colB = st.columns(2)
        with colA:
            st.caption("Top 5")
            st.table(df_cat.sort_values(category, ascending=False).head(5).reset_index(drop=True))
        with colB:
            st.caption("Bottom 5")
            st.table(df_cat.sort_values(category, ascending=True).head(5).reset_index(drop=True))

    with tab_top:
        left, right = st.columns([2, 1])
        with right:
            category_tb = st.selectbox("Category", options=available_categories, key="tb_cat")
            mode_rank = st.radio("Mode", ["Top", "Bottom"], horizontal=True)
            top_n = st.slider("Count", min_value=5, max_value=30, value=10, step=1)
            include_zero = st.checkbox("Include zero values", value=False)

        df_tb = template_df[[site_name_col, category_tb]].copy()
        df_tb[category_tb] = as_numeric(df_tb[category_tb])
        df_tb = df_tb.dropna(subset=[site_name_col])
        if not include_zero:
            df_tb = df_tb[df_tb[category_tb] != 0]
        df_tb = df_tb.sort_values(category_tb, ascending=(mode_rank == "Bottom")).head(top_n)

        with left:
            st.subheader(f"{mode_rank} {top_n} Companies — {category_tb}")
            st.dataframe(df_tb.reset_index(drop=True))

        st.markdown("### Visualization")
        if PLOTLY_AVAILABLE:
            fig = px.bar(
                df_tb.sort_values(category_tb, ascending=True),
                x=category_tb,
                y=site_name_col,
                orientation="h",
                title=f"{mode_rank} {top_n} — {category_tb}",
                template="simple_white",
                height=500,
            )
            fig.update_layout(margin=dict(t=60, r=20, b=40, l=80))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.bar_chart(df_tb.set_index(site_name_col), height=500, use_container_width=True)

        # Download
        csv = df_tb.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download Data (CSV)", data=csv, file_name="ranking.csv", mime="text/csv")

    with tab_matrix:
        st.subheader("Compare Multiple Categories (Matrix)")
        selected_cats = st.multiselect(
            "Select up to 6 categories",
            options=available_categories,
            default=available_categories[:3] if len(available_categories) >= 3 else available_categories,
            max_selections=6,
        )
        top_for_matrix = st.slider("Top N companies (by selected total)", 5, 30, 10)

        if selected_cats:
            df_m = template_df[[site_name_col] + selected_cats].copy()
            for c in selected_cats:
                df_m[c] = as_numeric(df_m[c])
            df_m["__total__"] = df_m[selected_cats].sum(axis=1)
            df_m = df_m.sort_values("__total__", ascending=False).head(top_for_matrix)
            df_show = df_m.drop(columns=["__total__"]).set_index(site_name_col)

            if PLOTLY_AVAILABLE:
                fig = px.imshow(
                    df_show,
                    color_continuous_scale="Blues",
                    aspect="auto",
                    title="Value Heatmap",
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.dataframe(df_show)
        else:
            st.info("Please select at least one category.")

    with tab_profile:
        st.subheader("Company Profile")
        companies = template_df[site_name_col].dropna().astype(str).unique().tolist()
        sel_company = st.selectbox("Select Company", options=companies)
        # Gather all available categories for the profile
        prof_cats = [c for c in available_categories if c in template_df.columns]
        df_p = template_df[[site_name_col] + prof_cats].copy()
        for c in prof_cats:
            df_p[c] = as_numeric(df_p[c])
        row = df_p[df_p[site_name_col].astype(str) == str(sel_company)]
        if row.empty:
            st.warning("Company data not found.")
        else:
            row_vals = row[prof_cats].iloc[0]
            cA, cB = st.columns([1, 1])
            with cA:
                st.markdown(f"**Partner Name:** {sel_company}")
                st.table(pd.DataFrame({"Category": prof_cats, "Value": row_vals.values}))
            with cB:
                if PLOTLY_AVAILABLE and len(prof_cats) >= 3:
                    # Radar chart
                    plot_df = pd.DataFrame({"Category": prof_cats, "Value": row_vals.values})
                    fig = px.line_polar(plot_df, r="Value", theta="Category", line_close=True, template="simple_white")
                    fig.update_traces(fill='toself')
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.bar_chart(row_vals, use_container_width=True)


def render_input():
    st.header("🧾 Data Input & Processing")
    st.write("This app counts occurrences from a source Excel file and writes them into a template file.")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("1. Upload Source File")
        source_file = st.file_uploader("Choose the Excel file containing raw data", type=["xlsx"], key="source_uploader")

    with col2:
        st.subheader("2. Upload Template File")
        template_file = st.file_uploader("Choose the target Excel template file", type=["xlsx", "xlsm"], key="template_uploader")

    if source_file and template_file:
        try:
            source_df = pd.read_excel(source_file, header=1)
            source_df.columns = source_df.columns.str.strip()
            template_df = pd.read_excel(template_file, sheet_name="Master Sheet")

            st.session_state.template_df = template_df

            st.subheader("3. Configure Processing Options")
            target_column = st.selectbox(
                "Select the target column in 'Master Sheet' to place the counts:",
                options=template_df.columns,
                key="target_column_select",
            )

            mode = st.radio(
                "Choose the update mode:",
                options=["Add (Tambah)", "Replace (Ganti)"],
                help="Add: add the new counts to existing values. Replace: overwrite existing values with the new counts.",
                key="mode_radio",
            )

            if st.button("🚀 Process Now!", key="process_button"):
                with st.spinner("Processing data..."):
                    result_df = process_data(source_df, template_df, target_column, mode)
                    st.session_state.result_df = result_df
                    st.subheader("4. Result")
                    st.write("Data processed successfully. Here is a preview of the result:")
                    st.dataframe(result_df.fillna(''))

                    # Tulis hasil ke workbook template untuk menjaga format
                    template_file.seek(0)
                    wb = openpyxl.load_workbook(template_file, keep_vba=True)
                    ws = wb["Master Sheet"]

                    header = [cell.value for cell in ws[1]]
                    site_name_col = header[0]
                    site_name_to_row = {}
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row=row, column=1).value
                        if val:
                            site_name_to_row[val] = row

                    for _, row_data in result_df.iterrows():
                        site_name = row_data[site_name_col]
                        if site_name in site_name_to_row:
                            row_idx = site_name_to_row[site_name]
                        else:
                            row_idx = ws.max_row + 1
                            ws.insert_rows(row_idx)
                            ws.cell(row=row_idx, column=1, value=site_name)
                            site_name_to_row[site_name] = row_idx
                        for col_idx, col_name in enumerate(header, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, None))

                    output = io.BytesIO()
                    wb.save(output)
                    processed_data = output.getvalue()

                    ext = ".xlsm" if template_file.name.lower().endswith(".xlsm") else ".xlsx"
                    st.session_state.last_processed_ext = ext
                    mime_type = (
                        "application/vnd.ms-excel.sheet.macroEnabled.12"
                        if ext == ".xlsm"
                        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Save to SQLite database for Dashboard auto-use
                    try:
                        run_meta = save_result_to_db(result_df, header)
                        if run_meta:
                            st.success(
                                f"Saved to local database '{DB_PATH.name}' (run id {run_meta['run_id']}). Dashboard will use this automatically.")
                    except Exception as db_err:
                        st.warning(f"Failed to save to database: {db_err}")

                    st.download_button(
                        label="📥 Download Result File",
                        data=processed_data,
                        file_name=f"processed_result{ext}",
                        mime=mime_type,
                        key="download_button",
                    )
        except Exception as e:
            st.error(f"An error occurred: {e}")
            st.warning("Make sure the uploaded Excel files are valid and the template contains a sheet named 'Master Sheet'.")
    else:
        st.info("Please upload both Excel files to begin.")


# --- Router ---
def render_guide():
    st.header("📘 User Guide")
    st.write("A comprehensive guide to using this application — focused on the deployed Streamlit Cloud version.")

    st.success("Use the cloud version: https://fillmastersheet.streamlit.app/ (click the link)")
    st.markdown("""
### Quick Start
1. Open: [fillmastersheet.streamlit.app](https://fillmastersheet.streamlit.app/)
2. Go to the "Data Input" menu.
3. Upload Source File (.xlsx) and Template File (.xlsx/.xlsm).
4. Choose the target column and mode (Add/Replace), then click "Process Now!".
5. Download the result (extension follows the template; .xlsm preserves macros).
6. Go to the "Dashboard" menu to explore data: Overview, Top/Bottom, Matrix, and Company Profile.
""")

    st.markdown("""
## 1. Application Summary
This app helps you:
- Count occurrences per company (based on the "Site Name" column) from the source Excel file.
- Update the template Excel file on the "Master Sheet" using the selected target column with two modes: Add and Replace.
- Explore the results on the Dashboard: Overview, Top/Bottom, Matrix comparisons, and Company Profile.
""")

    with st.expander("2. Data Preparation (Required)", expanded=False):
        st.markdown("""
### 2.1. Source File (Raw Data)
- Format: .xlsx
- Header is expected on the second row (this code uses `header=1`).
- Minimum required columns:
  - `Student Code` and `Course Code` (rows with missing values in either will be ignored)
  - `Site Name` (company/partner name)
- Other columns are ignored by the counting process but may be present.

### 2.2. Template File
- Format: .xlsx or .xlsm (macros are preserved because `keep_vba=True`).
- Must contain a sheet named `Master Sheet`.
- The first column (column A) should be the company name (e.g., `Partner Name / Site Name / Company Name`).
- Subsequent columns (B, C, etc.) are categories shown in the Dashboard. Non-numeric values (e.g. `Y`) are treated as 0 for visualization purposes.
- The header row is assumed to be the first row of the sheet.
""")

    with st.expander("3. Processing Mode: Add vs Replace"):
        st.markdown("""
- Add: the target column values will be incremented by the new counts.
  - Example: old value 10, new count 3 → stored 13.
- Replace: the target column values will be replaced by the new counts.
  - Example: old value 10, new count 3 → stored 3.
- If a company is not present in the template, a new row will be added automatically.
""")

    with st.expander("4. Steps in the Data Input Menu", expanded=True):
        st.markdown("""
1) Upload the Source File (.xlsx).
2) Upload the Template File (.xlsx or .xlsm) that contains a `Master Sheet`.
3) Select the target column (from the `Master Sheet` header) to receive the counts.
4) Choose the Mode (Add/Replace).
5) Click "Process Now!" and wait for the preview to appear.
6) Download the result using the Download button. Extension follows the template (if template is .xlsm the result will also be .xlsm and macros are preserved).

Technical notes when saving to the template:
- The app reads all headers from the first row of `Master Sheet`.
- It finds company rows by matching the first column.
- If not found, a new row is inserted at the end and all header columns are filled.
""")

    with st.expander("5. Steps in the Dashboard Menu"):
        st.markdown("""
- Upload (or reuse the one uploaded via Data Input) the template file to use as the Dashboard source.
- Categories shown in the Dashboard are taken automatically from the `Master Sheet` headers excluding the first column.

5.1 Overview
- KPI Cards: Total, Average, Maximum, Minimum.
- Value Distribution: histogram for the selected category.
- Sample Data: Top 5 and Bottom 5.

5.2 Top/Bottom
- Choose Top or Bottom mode, the number N, and whether to include zero values.
- Table and horizontal bar chart are available.
- Downloadable as CSV.

5.3 Matrix
- Select up to 6 categories to compare.
- Choose Top N companies based on the sum of selected categories.
- Displayed as a heatmap (or table if Plotly is not available).

5.4 Company Profile
- Select one company to view all available category values.
- Radar or bar chart (fallback) will be displayed.
""")

    with st.expander("6. Tips, Limitations, and Best Practices"):
        st.markdown("""
- Non-numeric values in category columns are treated as 0 for Dashboard visuals. Convert them first if you want them counted (e.g. `Y` → 1).
- Ensure exact column names (case-sensitive) for `Student Code`, `Course Code`, `Site Name` in the source file.
- For large files, keep them on local disk (not a network drive) for performance.
- Avoid duplicate company names in the template; if present, the first matching row will be updated.
- Keep a backup of the template before overwriting, especially for `.xlsm` files with important macros.
""")

    with st.expander("7. Troubleshooting (Common Issues)"):
        st.markdown("""
- "Sheet 'Master Sheet' not found": ensure the sheet name matches exactly, including case.
- "Column not found": check header spelling. For the source file, ensure `Student Code`, `Course Code`, `Site Name` exist.
- "Empty charts / all zeros": category column may contain non-numeric text or filters removed all rows.
- "Cannot download": check file size and browser download settings.
""")

    with st.expander("8. Notes for Streamlit Cloud Users", expanded=True):
        st.markdown("""
- Your data is only used in your session; avoid uploading sensitive data if unnecessary.
- Upload file size is subject to Streamlit Cloud limits. For large files consider:
  - Removing unnecessary sheets
  - Compressing or summarizing the source data
- If uploads are slow or fail, check your internet connection and retry.
- Recommended browsers: modern Chrome or Edge.
""")

    with st.expander("9. Run Locally (Optional)"):
        st.markdown("""
Optional commands to run locally on Windows PowerShell:
```powershell
# (Optional) Install dependencies for the project
# pip install -r requirements.txt

# Run the app
streamlit run "c:\\Users\\PRIMA\\OneDrive\\Documents\\PROJECT\\0 TRIAL\\Project Fill in Master sheet\\app.py"
```
""")

    with st.expander("10. FAQ"):
        st.markdown("""
- Do categories need to be defined manually? No. Categories are read automatically from the `Master Sheet` header (except the first column).
- Are macros removed when saving? No, `.xlsm` macros are preserved (`keep_vba=True`).
- Can I change the source header row? Currently the code expects the header on row 2 (`header=1`). Adjust code if different.
""")


if menu == "Dashboard":
    render_dashboard()
elif menu == "Data Input":
    render_input()
else:
    render_guide()

# --- Footer (fixed at bottom) ---
st.markdown(
    """
    <style>
        .footer {
            position: fixed;
            left: 0;
            right: 0;
            bottom: 0;
            width: 100%;
            background-color: #f0f2f6;
            color: #333;
            text-align: center;
            padding: 5px;
            font-size: 14px;
            border-top: 1px solid #ddd;
            z-index: 9999;
        }
        .footer p {
            margin-bottom: 2px;
            line-height: 1.2;
        }
        .footer a {
            color: #4b89ff;
            text-decoration: none;
            font-weight: bold;
        }
        .footer a:hover {
            text-decoration: underline;
        }
        /* Tambahkan padding bawah pada konten utama agar tidak tertutup footer */
        .main .block-container {
            padding-bottom: 48px;
        }
    </style>
    <div class="footer">
        <p>Developed by <b>Galih Primananda</b> </p>
        <p>
            <a href="https://instagram.com/glh_prima/" target="_blank">Instagram</a> |
            <a href="https://linkedin.com/in/galihprime/" target="_blank">LinkedIn</a> |
            <a href="https://github.com/PrimeFox59" target="_blank">GitHub</a> |
            <a href="https://drive.google.com/drive/folders/11ov7TpvOZ3m7k5GLRAbE2WZFbGVK2t7i?usp=sharing" target="_blank">Portfolio</a> |
            <a href="https://fastwork.id/user/glh_prima" target="_blank">Fastwork for Services & Collaboration</a>
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)
